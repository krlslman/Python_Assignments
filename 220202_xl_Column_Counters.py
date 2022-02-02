#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Feb  2 16:26:10 2022
@author: koral007
"""
from openpyxl import Workbook, load_workbook
import pandas as pd
# from collections import Counter

wb = load_workbook('/home/koral007/Downloads/row_count_sample.xlsx') 
ws = wb.active
file_name =   '/home/koral007/Downloads/row_count_sample.xlsx'
sheet = 'Sheet1'
df = pd.read_excel(io=file_name, sheet_name=sheet)

#number of non-empty cells in the column
print("dropna :  ", len(df['ID'].dropna()))
#no matter which column name your enter, returns the max one
print("notnull : ", df['ID'].notnull().count())
#returns the max one ????????????????
print("max_row : ", ws.max_row) 
#returns the max one ????????????????
print("len(ws['ID']) : ", len(ws['ID']))
#returns number of empty cells on column with index 0
print("empty cells : ", sum(df.apply(lambda x: x.isna().sum(), axis='columns')>0))
